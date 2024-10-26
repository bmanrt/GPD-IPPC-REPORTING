import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation  # Add this import
import sqlite3
import json
from datetime import datetime
from church_records import (
    CURRENCIES, 
    CONVERSION_RATES, 
    convert_to_espees, 
    handle_church_submit,
    handle_ror_outreaches_submit,
    handle_cell_submit
)

# Update the site name wherever it appears
SITE_NAME = "GPD Reporting"

# If there's a template header or similar, it might look like:
header_template = """
    <div class="header">
        <h1>GPD Reporting</h1>
        <!-- ... -->
    </div>
"""

def get_user_details(username):
    conn = sqlite3.connect('users.db')
    c = conn.cursor()
    c.execute("SELECT user_group, sub_group, region, zone FROM users WHERE username=?", (username,))
    user_details = c.fetchone()
    conn.close()
    return user_details

# Update the categories to include currency field
CATEGORIES = {
    "Adult Partner": [
        "Title",
        "First Name",
        "Surname",
        "Church", 
        "Group",
        "KingsChat Phone Number",
        "Email Address",
        "Currency",  # Added currency field
        "Total Amount Given for Wonder Challenge",
        "Total Amount Given for Rhapsody Languages",
        "Total Amount Given for Kiddies Products",
        "Total Amount Given for Teevo",
        "Total Amount Given for Braille(NOLB)",
        "Total Amount Given for Youth Aglow",
        "Total Given for Local Distribution",
        "Total Given for Subscriptions/Dubais"
    ],
    
    "Child Partner": [
        "Title",
        "First Name", 
        "Surname", 
        "Age",
        "KingsChat Phone Number", 
        "Email Address", 
        "Birthdays",
        "Church", 
        "Group",
        "Currency",  # Added currency field
        "Total Amount Given for Wonder Challenge",
        "Total Amount Given for Rhapsody Languages",
        "Total Amount Given for Kiddies Products",
        "Total Amount Given for Teevo",
        "Total Amount Given for Braille(NOLB)",
        "Total Amount Given for Youth Aglow",
        "Total Given for Local Distribution",
        "Total Given for Subscriptions/Dubais"
    ],
    
    "Teenager Partner": [
        "Title",
        "First Name", 
        "Surname",
        "KingsChat Phone Number", 
        "Email Address", 
        "Birthdays",
        "Church", 
        "Group",
        "Currency",  # Added currency field
        "Total Amount Given for Wonder Challenge",
        "Total Amount Given for Rhapsody Languages",
        "Total Amount Given for Kiddies Products",
        "Total Amount Given for Teevo",
        "Total Amount Given for Braille(NOLB)",
        "Total Amount Given for Youth Aglow",
        "Total Given for Local Distribution",
        "Total Given for Subscriptions/Dubais"
    ],
    
    "External Partner": [
        "Title",
        "First Name",
        "Surname",
        "KingsChat Phone Number",
        "Email Address",
        "Currency",  # Added currency field
        "Rhapsody Subscriptions and Dubais",
        "Sponsorship Through Retail Center",
        "Quantity Sponsored Through IRCON",
        "Translators Network International",
        "Rhapsody Influencers Network",
        "RIM",
        "Total"
    ],
    
    # Church Records Categories
    "Church Sponsorship": [
        "Zone Name",
        "Group Name",
        "Church Name",
        "Church Pastor",
        "KingsChat Phone Number",
        "Email Address",
        "Total Quantity Sponsored",
        "Currency",  # Added currency field
        "Total Amount",
        "Kiddies Products",
        "Teevo",
        "Braille(NOLB)",
        "Languages",
        "Youth Aglow"
    ],
    
    "Cell": [
        "Zone Name",
        "Cell Name",
        "Cell Leader",
        "KingsChat Phone Number",
        "Email Address",
        "Church",
        "Group",
        "Total Quantity Sponsored",
        "Currency",  # Added currency field
        "Total Amount Received",
        "Total Amount Given",
        "Kiddies Products",
        "Teevo",
        "Braille",
        "Languages",
        "Youth Aglow"
    ],
    
    # Separate ROR category
    "ROR Outreaches": [
        "Zone Name",
        "Group Name",
        "Reachout World Programs",
        "Rhapathon with Pastor Chris",
        "Reachout World Nations",
        "Say Yes to Kids",
        "Teevolution",
        "Youth Aglow",
        "No One Left Behind",
        "Penetrating with Truth",
        "Penetrating with Languages",
        "Adopt a Street",
        "Currency",  # Added currency field
        "Total Amount"
    ]
}

def create_blank_template():
    """Create a blank template with separate sheets for each category"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet
    
    # Style headers
    header_style = {
        'font': Font(bold=True),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'fill': PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    }
    
    # Remove Zone Name and Currency from all categories
    sheets_config = {
        "Adult Partners": [f for f in CATEGORIES["Adult Partner"] 
                         if f not in ["Zone Name", "Currency"]],
        "Child Partners": [f for f in CATEGORIES["Child Partner"] 
                         if f not in ["Zone Name", "Currency"]],
        "Teenager Partners": [f for f in CATEGORIES["Teenager Partner"] 
                            if f not in ["Zone Name", "Currency"]],
        "External Partners": [f for f in CATEGORIES["External Partner"] 
                            if f not in ["Zone Name", "Currency"]],
        "Category A Churches": [f for f in CATEGORIES["Church Sponsorship"] 
                              if f not in ["Zone Name", "Currency"]],
        "Category B Churches": [f for f in CATEGORIES["Church Sponsorship"] 
                              if f not in ["Zone Name", "Currency"]],
        "Churches": [f for f in CATEGORIES["Church Sponsorship"] 
                    if f not in ["Zone Name", "Currency"]],
        "Cell Records": [f for f in CATEGORIES["Cell"] 
                        if f not in ["Zone Name", "Currency"]],
        "ROR Outreaches": [f for f in CATEGORIES["ROR Outreaches"] 
                          if f not in ["Zone Name", "Currency"]]
    }
    
    # Create and format each sheet
    for sheet_name, fields in sheets_config.items():
        ws = wb.create_sheet(sheet_name)
        
        # Add headers
        for col, field in enumerate(fields, start=1):
            cell = ws.cell(row=1, column=col, value=field)
            for style_attr, style_value in header_style.items():
                setattr(cell, style_attr, style_value)
            
            # Set column width
            ws.column_dimensions[get_column_letter(col)].width = max(len(field) + 2, 15)
    
    # Add an Instructions sheet
    instructions = wb.create_sheet("Instructions", 0)
    instructions.cell(row=1, column=1, value="Instructions for Filling the Template")
    instructions.cell(row=3, column=1, value="1. Each category has its own sheet")
    instructions.cell(row=4, column=1, value="2. Fill in only the sheets you need")
    instructions.cell(row=5, column=1, value="3. Zone will be set automatically")
    instructions.cell(row=6, column=1, value="4. Currency will be selected when uploading")
    instructions.cell(row=7, column=1, value="5. All amount fields should be numbers only (no currency symbols)")
    instructions.cell(row=8, column=1, value="6. All quantity fields should be whole numbers")
    instructions.cell(row=9, column=1, value="7. Do not modify the column headers")
    instructions.cell(row=10, column=1, value="8. Do not add or remove columns")
    
    # Format instructions
    for row in range(1, 11):
        cell = instructions.cell(row=row, column=1)
        cell.font = Font(size=11)
        if row == 1:
            cell.font = Font(bold=True, size=12)
    
    instructions.column_dimensions['A'].width = 60
    
    return wb

def process_uploaded_records(df, selected_zone):
    """Process uploaded records with improved validation"""
    if isinstance(df, dict):  # Excel file with multiple sheets
        # Show preview of data first
        has_data = False
        for sheet_name, sheet_df in df.items():
            if not sheet_df.empty:
                has_data = True
                st.subheader(f"Preview of {sheet_name}")
                st.dataframe(sheet_df.head(), use_container_width=True)
        
        if not has_data:
            st.warning("No data found in the uploaded file.")
            return
        
        # Single currency selection for all records
        currency = st.selectbox(
            "Select Currency for All Records",
            CURRENCIES,
            key="global_currency"
        )
        
        valid_sheets = {
            'partners': [],
            'church': [],
            'ror': []
        }
        
        # Process partner sheets with amount calculations
        partner_sheets = {
            "Adult Partners": ("Adult Partner", "partners"),
            "Child Partners": ("Child Partner", "partners"),
            "Teenager Partners": ("Teenager Partner", "partners"),
            "External Partners": ("External Partner", "partners")
        }
        
        # Process church sheets
        church_sheets = {
            "Category A Churches": ("Category A", "church"),
            "Category B Churches": ("Category B", "church"),
            "Churches": ("Church", "church"),
            "Cell Records": ("Cell", "church")
        }
        
        # Process each sheet
        if st.button("Validate and Submit Records"):
            for sheet_name, sheet_df in df.items():
                if sheet_df.empty:
                    continue
                
                # Add zone and currency to the data
                sheet_df['Zone'] = selected_zone
                sheet_df['Currency'] = currency
                
                if sheet_name in partner_sheets:
                    category, sheet_type = partner_sheets[sheet_name]
                    
                    # Calculate total amount for partners
                    if category != "External Partner":
                        sheet_df['Total Amount'] = sheet_df[[
                            'Total Amount Given for Wonder Challenge',
                            'Total Amount Given for Rhapsody Languages',
                            'Total Amount Given for Kiddies Products',
                            'Total Amount Given for Teevo',
                            'Total Amount Given for Braille(NOLB)',
                            'Total Amount Given for Youth Aglow',
                            'Total Given for Local Distribution',
                            'Total Given for Subscriptions/Dubais'
                        ]].fillna(0).sum(axis=1)
                    else:
                        sheet_df['Total Amount'] = sheet_df[[
                            'Rhapsody Subscriptions and Dubais',
                            'Sponsorship Through Retail Center',
                            'Translators Network International',
                            'Rhapsody Influencers Network',
                            'RIM'
                        ]].fillna(0).sum(axis=1)

                    if validate_sheet_data(sheet_df, CATEGORIES[category]):
                        sheet_df['Category'] = category
                        valid_sheets[sheet_type].append(sheet_df)
                        st.success(f"✓ {sheet_name} data validated successfully")
                    else:
                        st.error(f"✗ {sheet_name} data validation failed")
                        
                elif sheet_name in church_sheets:
                    category, sheet_type = church_sheets[sheet_name]
                    if validate_sheet_data(sheet_df, CATEGORIES["Church Sponsorship"]):
                        sheet_df['Category'] = category
                        valid_sheets[sheet_type].append(sheet_df)
                        st.success(f"✓ {sheet_name} data validated successfully")
                    else:
                        st.error(f"✗ {sheet_name} data validation failed")
                        
                elif sheet_name == "ROR Outreaches":
                    if validate_sheet_data(sheet_df, CATEGORIES["ROR Outreaches"]):
                        valid_sheets['ror'].append(sheet_df)
                        st.success("✓ ROR Outreaches data validated successfully")
                    else:
                        st.error("✗ ROR Outreaches data validation failed")
            
            # Save validated data
            success_count = 0
            error_count = 0
            
            # Save partner records with calculated totals
            if valid_sheets['partners']:
                partner_df = pd.concat(valid_sheets['partners'], ignore_index=True)
                if save_partner_records(partner_df, selected_zone, currency):
                    success_count += 1
                    st.success("✓ Partner records saved successfully!")
                else:
                    error_count += 1
                    st.error("✗ Error saving partner records")
            
            # Save church records
            if valid_sheets['church']:
                church_df = pd.concat(valid_sheets['church'], ignore_index=True)
                if save_church_records(church_df, selected_zone, currency):
                    success_count += 1
                    st.success("✓ Church records saved successfully!")
                else:
                    error_count += 1
                    st.error("✗ Error saving church records")
            
            # Save ROR records
            if valid_sheets['ror']:
                ror_df = pd.concat(valid_sheets['ror'], ignore_index=True)
                if save_ror_records(ror_df, selected_zone, currency):
                    success_count += 1
                    st.success("✓ ROR records saved successfully!")
                else:
                    error_count += 1
                    st.error("✗ Error saving ROR records")
            
            if error_count == 0 and success_count > 0:
                st.success(f"All {success_count} record types submitted successfully!")
                if st.button("Upload More Records"):
                    st.experimental_rerun()
            elif error_count > 0:
                st.warning(f"Completed with {success_count} successes and {error_count} errors.")
    else:
        st.error("Please use the Excel template. CSV uploads are not supported.")

def validate_sheet_data(df, required_fields):
    """Validate sheet data against required fields, with more flexible validation"""
    try:
        # Determine category based on the sheet's required fields
        category_mapping = {
            frozenset(['First Name', 'Surname', 'Church', 'Group']): 'Adult Partner',
            frozenset(['First Name', 'Surname', 'Age', 'Church', 'Group']): 'Child Partner',
            frozenset(['First Name', 'Surname', 'Church', 'Group', 'Birthdays']): 'Teenager Partner',
            frozenset(['First Name', 'Surname', 'Rhapsody Subscriptions and Dubais']): 'External Partner',
            frozenset(['Church Name', 'Group Name', 'Church Pastor']): 'Church Sponsorship',
            frozenset(['Cell Name', 'Cell Leader', 'Church', 'Group']): 'Cell',
            frozenset(['Group Name', 'Reachout World Programs']): 'ROR Outreaches'
        }

        # Get the set of fields present in the DataFrame
        df_fields = set(df.columns)
        
        # Find matching category
        category = None
        for fields, cat in category_mapping.items():
            if fields.issubset(df_fields):
                category = cat
                break
        
        if not category:
            st.error(f"Could not determine record type from columns: {', '.join(df_fields)}")
            return False

        # Validate numeric fields based on category
        numeric_fields = []
        if category in ['Adult Partner', 'Child Partner', 'Teenager Partner']:
            numeric_fields = [
                'Total Amount Given for Wonder Challenge',
                'Total Amount Given for Rhapsody Languages',
                'Total Amount Given for Kiddies Products',
                'Total Amount Given for Teevo',
                'Total Amount Given for Braille(NOLB)',
                'Total Amount Given for Youth Aglow',
                'Total Given for Local Distribution',
                'Total Given for Subscriptions/Dubais'
            ]
        elif category == 'External Partner':
            numeric_fields = [
                'Rhapsody Subscriptions and Dubais',
                'Sponsorship Through Retail Center',
                'Quantity Sponsored Through IRCON',
                'Translators Network International',
                'Rhapsody Influencers Network',
                'RIM',
                'Total'
            ]
        elif category == 'Church Sponsorship':
            numeric_fields = [
                'Total Quantity Sponsored',
                'Total Amount',
                'Kiddies Products',
                'Teevo',
                'Braille(NOLB)',
                'Languages',
                'Youth Aglow'
            ]
        elif category == 'Cell':
            numeric_fields = [
                'Total Quantity Sponsored',
                'Total Amount Received',
                'Total Amount Given',
                'Kiddies Products',
                'Teevo',
                'Braille',
                'Languages',
                'Youth Aglow'
            ]
        elif category == 'ROR Outreaches':
            numeric_fields = [
                'Reachout World Programs',
                'Rhapathon with Pastor Chris',
                'Reachout World Nations',
                'Say Yes to Kids',
                'Teevolution',
                'Youth Aglow',
                'No One Left Behind',
                'Penetrating with Truth',
                'Penetrating with Languages',
                'Adopt a Street',
                'Total Amount'
            ]

        # Validate numeric fields
        for field in numeric_fields:
            if field in df.columns:
                try:
                    # Replace empty or non-numeric values with 0
                    df[field] = pd.to_numeric(df[field].fillna(0))
                except Exception as e:
                    st.error(f"Non-numeric values found in field: {field}")
                    return False

        return True
        
    except Exception as e:
        st.error(f"Validation error: {str(e)}")
        return False

def record_templates_ui():
    st.title("Record Templates and Upload")
    
    if st.session_state.is_super_admin:
        # Load zones data
        with open('zones_data.json', 'r') as f:
            zones_data = json.load(f)
        
        # First select region
        regions = list(zones_data.keys())
        selected_region = st.selectbox("Select Region", regions)
        
        # Then select zone from that region
        if selected_region:
            zones = zones_data[selected_region]
            selected_zone = st.selectbox("Select Zone", zones)
    else:
        # For RZMs, get their zone from user details
        user_details = get_user_details(st.session_state.username)
        if user_details:
            _, _, _, selected_zone = user_details
            if selected_zone:
                st.info(f"Uploading records for: {selected_zone}")
            else:
                st.error("No zone assigned to your account")
                return
        else:
            st.error("Could not retrieve user details")
            return
    
    tab1, tab2 = st.tabs(["Download Template", "Upload Records"])
    
    with tab1:
        st.header("Download Blank Template")
        if st.button("Download Template"):
            buffer = BytesIO()
            wb = create_blank_template()
            wb.save(buffer)
            buffer.seek(0)
            
            st.download_button(
                label="Click here to download",
                data=buffer,
                file_name="blank_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    with tab2:
        st.header("Upload Records")
        if not selected_zone:
            st.warning("Please select a zone before uploading records.")
            return
        
        uploaded_file = st.file_uploader("Choose Excel file", type=["xlsx"])
        if uploaded_file is not None:
            try:
                df = pd.read_excel(uploaded_file, sheet_name=None)
                process_uploaded_records(df, selected_zone)
            except Exception as e:
                st.error(f"Error reading file: {e}")

# Add these functions after the validate_sheet_data function

def save_partner_records(df, selected_zone, currency):
    """Save partner records to database"""
    try:
        from partner_records import add_partner_record, add_external_partner_record
        
        for _, row in df.iterrows():
            partner_data = row.to_dict()
            partner_data['zone'] = selected_zone
            partner_data['currency'] = currency
            
            # Ensure total amount is included
            if 'Total Amount' not in partner_data:
                if row['Category'] != 'External Partner':
                    partner_data['total_amount'] = sum([
                        float(partner_data.get('Total Amount Given for Wonder Challenge', 0)),
                        float(partner_data.get('Total Amount Given for Rhapsody Languages', 0)),
                        float(partner_data.get('Total Amount Given for Kiddies Products', 0)),
                        float(partner_data.get('Total Amount Given for Teevo', 0)),
                        float(partner_data.get('Total Amount Given for Braille(NOLB)', 0)),
                        float(partner_data.get('Total Amount Given for Youth Aglow', 0)),
                        float(partner_data.get('Total Given for Local Distribution', 0)),
                        float(partner_data.get('Total Given for Subscriptions/Dubais', 0))
                    ])
                else:
                    partner_data['total_amount'] = sum([
                        float(partner_data.get('Rhapsody Subscriptions and Dubais', 0)),
                        float(partner_data.get('Sponsorship Through Retail Center', 0)),
                        float(partner_data.get('Translators Network International', 0)),
                        float(partner_data.get('Rhapsody Influencers Network', 0)),
                        float(partner_data.get('RIM', 0))
                    ])
            
            if row['Category'] == 'External Partner':
                success, message = add_external_partner_record(partner_data)
            else:
                success, message = add_partner_record(
                    partner_data,
                    is_child=(row['Category'] == 'Child Partner'),
                    is_teenager=(row['Category'] == 'Teenager Partner')
                )
            
            if not success:
                st.error(message)
                return False
        
        return True
    except Exception as e:
        st.error(f"Error saving partner records: {e}")
        return False

def save_church_records(df, selected_zone, currency):
    """Save church records to database"""
    try:
        from church_records import handle_church_submit, handle_cell_submit
        
        for _, row in df.iterrows():
            # Helper function to safely convert to int
            def safe_int(value, default=0):
                try:
                    if pd.isna(value):
                        return default
                    return int(float(value))
                except (ValueError, TypeError):
                    return default
            
            # Calculate total amount from quantities
            quantities = {
                'kiddies_products': safe_int(row.get('Kiddies Products')),
                'teevo': safe_int(row.get('Teevo')),
                'braille_nolb': safe_int(row.get('Braille(NOLB)')),
                'languages': safe_int(row.get('Languages')),
                'youth_aglow': safe_int(row.get('Youth Aglow'))
            }
            
            total_amount = sum(quantities.values())
            
            record_data = {
                'zone_name': selected_zone,
                'group_name': str(row.get('Group Name', '')),
                'church_name': str(row.get('Church Name', '')),
                'church_pastor': str(row.get('Church Pastor', '')),
                'kingschat_phone': str(row.get('KingsChat Phone Number', '')),
                'email': str(row.get('Email Address', '')),
                'total_quantity': safe_int(row.get('Total Quantity Sponsored')),
                'currency': currency,
                'total_amount': total_amount,
                **quantities
            }
            
            if row['Category'] in ['Category A', 'Category B', 'Church']:
                success, message = handle_church_submit(record_data, row['Category'][-1])
            elif row['Category'] == 'Cell':
                # For cell records, calculate received and given amounts
                cell_quantities = {
                    'total_quantity': safe_int(row.get('Total Quantity Sponsored')),
                    'kiddies_products': safe_int(row.get('Kiddies Products')),
                    'teevo': safe_int(row.get('Teevo')),
                    'braille': safe_int(row.get('Braille')),  # Note: Different field name for cells
                    'languages': safe_int(row.get('Languages')),
                    'youth_aglow': safe_int(row.get('Youth Aglow'))
                }
                
                total_amount = sum(cell_quantities.values())
                
                cell_data = {
                    'zone_name': selected_zone,
                    'cell_name': str(row.get('Cell Name', '')),
                    'cell_leader': str(row.get('Cell Leader', '')),
                    'kingschat_phone': str(row.get('KingsChat Phone Number', '')),
                    'email': str(row.get('Email Address', '')),
                    'church': str(row.get('Church', '')),
                    'group': str(row.get('Group', '')),
                    'currency': currency,
                    'total_amount_received': total_amount,
                    'total_amount_given': total_amount,
                    **cell_quantities
                }
                success, message = handle_cell_submit(cell_data)
            
            if not success:
                st.error(message)
                return False
        
        return True
    except Exception as e:
        st.error(f"Error saving church records: {e}")
        return False

def save_ror_records(df, selected_zone, currency):
    """Save ROR outreach records to database"""
    try:
        from church_records import handle_ror_outreaches_submit
        
        for _, row in df.iterrows():
            record_data = {
                "zone_name": selected_zone,
                "group_name": row.get('Group Name', ''),
                "reachout_world_programs": int(row.get('Reachout World Programs', 0)),
                "rhapathon": int(row.get('Rhapathon with Pastor Chris', 0)),
                "reachout_world_nations": int(row.get('Reachout World Nations', 0)),
                "say_yes_to_kids": int(row.get('Say Yes to Kids', 0)),
                "teevolution": int(row.get('Teevolution', 0)),
                "youth_aglow": int(row.get('Youth Aglow', 0)),
                "no_one_left_behind": int(row.get('No One Left Behind', 0)),
                "penetrating_truth": int(row.get('Penetrating with Truth', 0)),
                "penetrating_languages": int(row.get('Penetrating with Languages', 0)),
                "adopt_a_street": int(row.get('Adopt a Street', 0)),
                "currency": currency,
                "total_amount": float(row.get('Total Amount', 0))
            }
            
            # Calculate total outreaches
            record_data["total_outreaches"] = sum([
                record_data["reachout_world_programs"],
                record_data["rhapathon"],
                record_data["reachout_world_nations"],
                record_data["say_yes_to_kids"],
                record_data["teevolution"],
                record_data["youth_aglow"],
                record_data["no_one_left_behind"],
                record_data["penetrating_truth"],
                record_data["penetrating_languages"],
                record_data["adopt_a_street"]
            ])
            
            success, message = handle_ror_outreaches_submit(record_data)
            if not success:
                st.error(message)
                return False
        
        return True
    except Exception as e:
        st.error(f"Error saving ROR records: {e}")
        return False

# Add these constants near the top of the file
REQUIRED_FIELDS = {
    "Adult Partner": [
        "First Name",
        "Surname",
        "Church",
        "Group"
    ],
    
    "Child Partner": [
        "First Name",
        "Surname",
        "Age",
        "Church",
        "Group"
    ],
    
    "Teenager Partner": [
        "First Name",
        "Surname",
        "Church",
        "Group"
    ],
    
    "External Partner": [
        "First Name",
        "Surname"
    ],
    
    "Church Sponsorship": [
        "Church Name",
        "Group Name",
        "Total Amount"
    ],
    
    "Cell": [
        "Cell Name",
        "Cell Leader",
        "Church",
        "Group"
    ],
    
    "ROR Outreaches": [
        "Group Name",
        "Total Amount"
    ]
}

if __name__ == "__main__":
    record_templates_ui()

